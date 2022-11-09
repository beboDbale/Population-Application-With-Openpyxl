import openpyxl
import sys
import time as ti
from openpyxl.utils import get_column_letter

name = ""


def check_file():
    global wb
    try:
        wb = openpyxl.load_workbook(input("Enter the name of file or the path : ").strip())
        wb.active
    except:
        print("No file with tis name or path ")
        check_file()


def check_country():
    # to check the name of country and it is important for other functions
    global name
    name = input("Enter name of country : ").strip()
    found = False
    # while name.upper() not in wb.sheetnames and name.lower() not in wb.sheetnames.lower() and name.title() not in wb.sheetnames:
    # while name.lower() not in wb.sheetnames.lower():
    for i in wb.sheetnames:
        if name.lower() == i.lower():
            found = True
            break
    if not found:
        print("Invalid , these are the available countries : " + str(wb.sheetnames))
        check_country()
    return True


def name_of_city(name):
    # to display city's info by their names
    for i in wb.sheetnames:
        if name.lower() == i.lower():
            sh1 = wb[i]
            break
    found = False
    sheet_max_row = sh1.max_row
    name_city = input("\nenter name of city : ").strip()
    for row in range(1, sheet_max_row + 1):
        for col in range(1, 3):
            city = sh1['A' + str(row)].value
            char = get_column_letter(col)
            if name_city.lower() == city.lower():
                print(sh1[char + str(row)].value, end='   ')
                found = True
    num = 1
    if not found:
        print("Not available , these are available :")
        for row in range(1, sheet_max_row + 1):
            for col in range(1, 2):
                city = sh1['A' + str(row)].value
                char = get_column_letter(col)
                print(sh1[char + str(row)].value, end='   ')
                if row == (num * 10):
                    print()
                    num += 1
        name_of_city(name)


def all_population(name):
    # to display all country's population and total of people
    for i in wb.sheetnames:
        if name.lower() == i.lower():
            sh1 = wb[i]
            break
    sheet_max_row = sh1.max_row
    total = 0
    for row in range(1, sheet_max_row + 1):
        total += sh1['B' + str(row)].value
        for col in range(1, 3):
            char = get_column_letter(col)
            print(str(sh1[char + str(row)].value), end=" " * (40 - len(sh1["A" + str(row)].value)))
        print()
    print("Total  population in " + name + " is = " + str(total) + "\n")


def maximum_and_minimum(name):
    # to get maximum and minimum of one country
    max_pop = 0
    mini_pop = 100000000
    maximum_country = ""
    minimum_country = ""
    for i in wb.sheetnames:
        if name.lower() == i.lower():
            sh1 = wb[i]
            break
    sheet_max_row = sh1.max_row
    for row in range(1, sheet_max_row + 1):
        if max_pop < sh1['B' + str(row)].value:
            max_pop = sh1['B' + str(row)].value
            maximum_country = sh1['A' + str(row)].value
        if mini_pop > sh1['B' + str(row)].value:
            mini_pop = sh1['B' + str(row)].value
            minimum_country = sh1['A' + str(row)].value
    print("Maximum population in " + name + " is " + maximum_country + " with population of : " + str(max_pop))
    print("Minimum population in " + name + " is " + minimum_country + " with population of : " + str(mini_pop) + "\n")


check_file()
# (pyxl.xlsx)(test.xlsx) recommended files to open
# better to put the full path
while True:
    # infinity loop for user until he enter exit
    print("this application gives you some information and display it if it is available ^-^ ")
    if check_country():
        print("there is some choices : \n1 -Display the population of each state / province / governorate and "
              "total population of the country \n2- Display highest population and lowest one \n3- Display one city's "
              "information \n4- Exit")
        msg = input("Enter number of your choice please : ").strip()
    if msg == "1":
        all_population(name)
    elif msg == "2":
        maximum_and_minimum(name)
    elif msg == "3":
        name_of_city(name)
    elif msg == "4" or msg.lower() == 'exit':
        sys.exit()
    else:
        print("Wrong index!! , Try again \n \n")
    ti.sleep(4)
